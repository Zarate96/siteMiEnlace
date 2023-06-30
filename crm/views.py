import csv  
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.contrib.auth.mixins import UserPassesTestMixin

from .db import *

db = DatabaseConnection(
    settings.DB_CRM_HOST, 
    settings.DB_CRM_NAME,
    settings.DB_CRM_USER, 
    settings.DB_CRM_PASSWORD
)

class Crm(UserPassesTestMixin, TemplateView):
    template_name = 'crm/home.html'

    def test_func(self):
        return self.request.user.is_superuser
        
    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args,**kwargs)
        context['title'] = 'Home'
        return context

class Usuarios(UserPassesTestMixin, TemplateView):
    template_name = 'crm/usuarios.html'

    def test_func(self):
        return self.request.user.is_superuser
        
    def get_context_data(self, *args, **kwargs):
        query = db.excute_query("select id, userName, activo, nombre, apellidopaterno, apellidomaterno, email,isPersonaFisica, idEstado, ciudad, lada, celular from enlace.usuarios")
        db.disconnect()
        context = super().get_context_data(*args,**kwargs)
        context['title'] = 'Usuarios'
        context['usuarios'] = query
        return context

class Test(UserPassesTestMixin, TemplateView):
    template_name = 'crm/templateTest.html'

    def test_func(self):
        return self.request.user.is_superuser
        
    def get_context_data(self, *args, **kwargs):
        query = db.excute_query("""select  CONCAT( 'OCC',SOL.FOLIO) OCC_FOLIO, USR.username Proveedor,
                                    SOD.partida,
                                    COD.descripcion, COD.marca, COD.empaque, COD.cantidad, COD.precioUnitario, COD.precioExtendido, COD.iva
                                    from enlace.solicitudes SOL 
                                    inner join enlace.solicituddetalle SOD on SOL.id = SOD.idSolicitud
                                    inner join enlace.cotizaciondetalle COD on COD.ID = SOD.idCotizacionDetalle
                                    left JOIN enlace.USUARIOS USR ON USR.ID = COD.IDUSUARIO
                                    WHERE  SOL.status = 'confirmada'
                                    and CONCAT('OCC' , SOL.folio) = 'OCCSANATORIO0034';""")
        query2 = db.excute_query("""select * from enlace.usuarios 
                                    WHERE id='189';""")
        query3 = db.excute_query("""select * from enlace.datosfiscales
                                    WHERE idUsuario='189';""")
        db.disconnect()
        context = super().get_context_data(*args,**kwargs)
        context['title'] = 'Test'
        context['result'] = query
        context['result2'] = query2
        context['result3'] = query3
        return context

class Solicitudes(UserPassesTestMixin, TemplateView):
    template_name = 'crm/solicitudes.html'

    def test_func(self):
        return self.request.user.is_superuser
        
    def get_context_data(self, *args, **kwargs):
        query = db.excute_query("""select  * from enlace.solicitudes
                                    order by solicitudes.fecAlta DESC;""")
        db.disconnect()
        context = super().get_context_data(*args,**kwargs)
        context['title'] = 'Solicitudes'
        context['myresult'] = query
        return context
    
def oce_detalle(request):
    if request.method == 'POST':
        occ = request.POST['occ']
        oce = request.POST['oce']
        idProveedor = request.POST['idProveedor']

        try:
            #OCETAPERULE00140 OCCSANATORIO0036 154 160
            query = db.excute_query(f"""select CONCAT('OCC' , SOL.folio) OCC,
                                        (select CONCAT( 'OCE', CO.FOLIO)  from enlace.cotizacion CO where CONCAT( 'OCE', CO.FOLIO) = '{oce}') OCE,
                                        SOD.partida, COD.descripcion, COD.marca, COD.empaque, COD.cantidad, COD.precioUnitario, COD.precioExtendido, COD.iva
                                        from enlace.solicitudes SOL 
                                        inner join enlace.solicituddetalle SOD on SOL.id = SOD.idSolicitud
                                        inner join enlace.cotizaciondetalle COD on COD.ID = SOD.idCotizacionDetalle
                                        inner JOIN enlace.USUARIOS USR ON USR.ID = COD.IDUSUARIO
                                        WHERE  USR.id in (select idUsuario from enlace.cotizacion CO where CONCAT( 'OCE', CO.FOLIO) = '{oce}') 
                                        and CONCAT('OCC' , SOL.folio) = '{occ}';""")
            # query2 = db.excute_query(f"""select * from enlace.usuarios 
            #                             WHERE id='{idCliente}';""")
            query2 = db.excute_query(f"""select * from enlace.datosfiscales
                                        WHERE idUsuario='{idProveedor}';""")
            print(query2)
            # PERSONA FISICA  isPersonaFisica 1=Si, 0=No
            query3 = db.excute_query(f"""select isPersonaFisica from enlace.usuarios WHERE id='{idProveedor}';""")
            print(query3)
            # ESTADO
            idEsatdo = query2[0][11]
            print(idEsatdo)
            estado = db.excute_query(f"""select estado from enlace.estados WHERE id='{idEsatdo}';""")

            db.disconnect()
            if query and query2 and query3:
                subtotal_oce = 0
                iva_oce = 0
                for row in query:
                    subtotal_oce = row[8] + subtotal_oce
                for row in query:
                    iva_oce = row[9] + iva_oce
                total_oce = subtotal_oce + iva_oce
                isFisica = query3[0][0]
                name = f'{query2[0][3]} {query2[0][4]} {query2[0][5]}' if isFisica else f'{query2[0][2]}'
                address = f'Calle {query2[0][6]} #{query2[0][7]} {query2[0][8]},  Colonia {query2[0][9]} <br> {query2[0][10]}, {estado[0][0]} {query2[0][12]} <br> RFC: <b>{query2[0][13]}</b>'
                now = datetime.datetime.now()
                date = now.strftime('%d-%m-%Y')
            
            if not query:
                messages.success(request, f"Data no encontrada con OCE: '{oce}' y OCC: '{occ}'.")
            if not query2:
                messages.success(request, f"Proveedor '{idProveedor}' no encontrado.")
            
            context = {
                "title":"OCE Detalle",
                "result":query,
                "proveedor":query2,
                "date":date,
                "name":name,
                "address":address,
                "occ":occ,
                "oce":oce,
                "subtotal_oce":subtotal_oce,
                "iva_oce":iva_oce,
                "total_oce":total_oce,
            }

            return render(request, 'crm/oce_detalle.html', context)
        
        except Exception as e:
            messages.error(request, f'Ha occurrido un error: {e}')

    context = {
        "title":"OCE Detalle",
    }

    return render(request, 'crm/oce_detalle.html', context)

def costo_envio_occ(request):
    query = ""
    if request.method == 'POST':
        occ = request.POST['occ']
        try:
            query = db.excute_query(f"""select CONCAT( 'OCC',SOL.FOLIO) OCC_FOLIO, FL.alto, FL.ancho, FL.largo, FL.peso, FL.costo, FL.combustible, FL.IVA from enlace.fletesolicitud FL
                                inner join enlace.solicitudes SOL on SOL.id = FL.idSolicitud
                                where CONCAT('OCC' , SOL.folio) = '{occ}';'""")
            db.disconnect()
            if not query:
                messages.success(request, f"OCC '{occ}' no encontrada.")
        except:
            messages.error(request, 'Ha occurrido un error')

    context = {
        "title":"Costo Envío OCC",
        "result":query,
    }
    return render(request, 'crm/costo_envio_occ.html', context)

def facturacion_cliente(request):
    #OCCSANATORIO0034
    #189
    if request.method == 'POST':
        occ = request.POST['occ']
        idUsuario = request.POST['idUsuario']
        try:
            query = db.excute_query(f"""select  CONCAT( 'OCC',SOL.FOLIO) OCC_FOLIO, USR.username Proveedor,
                                    SOD.partida,
                                    COD.descripcion, COD.marca, COD.empaque, COD.cantidad, COD.precioUnitario, COD.precioExtendido, COD.iva
                                    from enlace.solicitudes SOL 
                                    inner join enlace.solicituddetalle SOD on SOL.id = SOD.idSolicitud
                                    inner join enlace.cotizaciondetalle COD on COD.ID = SOD.idCotizacionDetalle
                                    left JOIN enlace.USUARIOS USR ON USR.ID = COD.IDUSUARIO
                                    WHERE  SOL.status = 'confirmada'
                                    and CONCAT('OCC' , SOL.folio) = '{occ}';""")
            query2 = db.excute_query(f"""select * from enlace.usuarios 
                                        WHERE id='{idUsuario}';""")
            query3 = db.excute_query(f"""select * from enlace.datosfiscales
                                        WHERE idUsuario='{idUsuario}';""")
           
            # PERSONA FISICA  isPersonaFisica 1=Si, 0=No
            query4 = db.excute_query(f"""select isPersonaFisica from enlace.usuarios WHERE id='{idUsuario}';;""")

            query5 = db.excute_query(f"""select CONCAT( 'OCC',SOL.FOLIO) OCC_FOLIO, FL.alto, FL.ancho, FL.largo, FL.peso, FL.costo, FL.combustible, FL.IVA from enlace.fletesolicitud FL
                                inner join enlace.solicitudes SOL on SOL.id = FL.idSolicitud
                                where CONCAT('OCC' , SOL.folio) = '{occ}';'""")
            
            # ESTADO
            idEsatdo = query3[0][11]            
            estado = db.excute_query(f"""select estado from enlace.estados WHERE id='{idEsatdo}';""")

            db.disconnect()
            
            if query and query2 and query3:
                subtotal = 0
                isFisica = query4[0][0]
                name = f'{query3[0][3]} {query3[0][4]} {query3[0][5]}' if isFisica else f'{query3[0][2]}'
                address = f'Calle {query3[0][6]} #{query3[0][7]} {query3[0][8]},  Colonia {query3[0][9]} <br> {query3[0][10]}, {estado[0][0]} {query3[0][12]} <br> RFC: <b>{query3[0][13]}</b>'
                costo_envio = query5[0][6]
                iva_envio = query5[0][7]
                for row in query:
                    subtotal = row[8] + subtotal
                subtotal = float(subtotal) + float(costo_envio)
                iva = float(subtotal) * 0.16
                total = float(subtotal) + iva
                now = datetime.datetime.now()
                date = now.strftime('%d-%m-%Y')
                context = {
                    "title":"Factura Cliente",
                    "result":query,
                    "name":name,
                    "address":address,
                    "subtotal":round(subtotal, 2),
                    "iva":round(iva, 2),
                    "total":round(total, 2),
                    "costo_envio":costo_envio,
                    "occ":occ,
                    "idUsuario":idUsuario,
                    "query5":query5,
                    "date":date,
                }

            if not query:
                messages.success(request, f"OCC '{occ}' no encontrada.")
            if not query2:
                messages.success(request, f"Usuario '{idUsuario}' no encontrado.")

            return render(request, 'crm/facturacion_cliente.html', context)
        
        except Exception as e:
            messages.error(request, f'Ha occurrido un error: {e}')
            
    context = {"title":"Factura Cliente"}
    return render(request, 'crm/facturacion_cliente.html', context)

def facturacion_proveedor(request):
    #OCCSANATORIO0034
    #189
    if request.method == 'POST':
        occ = request.POST['occ']
        oce = request.POST['oce']
        idUsuario = request.POST['idUsuario']
        partida2 = request.POST['costo_envio']
        comision_original = request.POST['comision']
        comision = float(comision_original) * 0.01
        try:
            query = db.excute_query(f"""select CONCAT('OCC' , SOL.folio) OCC,
                                        (select CONCAT( 'OCE', CO.FOLIO)  from enlace.cotizacion CO where CONCAT( 'OCE', CO.FOLIO) = '{oce}') OCE,
                                        SOD.partida, COD.descripcion, COD.marca, COD.empaque, COD.cantidad, COD.precioUnitario, COD.precioExtendido, COD.iva
                                        from enlace.solicitudes SOL 
                                        inner join enlace.solicituddetalle SOD on SOL.id = SOD.idSolicitud
                                        inner join enlace.cotizaciondetalle COD on COD.ID = SOD.idCotizacionDetalle
                                        inner JOIN enlace.USUARIOS USR ON USR.ID = COD.IDUSUARIO
                                        WHERE  USR.id in (select idUsuario from enlace.cotizacion CO where CONCAT( 'OCE', CO.FOLIO) = '{oce}') 
                                        and CONCAT('OCC' , SOL.folio) = '{occ}';""")
            query2 = db.excute_query(f"""select * from enlace.usuarios 
                                        WHERE id='{idUsuario}';""")
            query3 = db.excute_query(f"""select * from enlace.datosfiscales
                                        WHERE idUsuario='{idUsuario}';""")
           
            # PERSONA FISICA  isPersonaFisica 1=Si, 0=No
            query4 = db.excute_query(f"""select isPersonaFisica from enlace.usuarios WHERE id='{idUsuario}';;""")
            
            # ESTADO
            idEsatdo = query3[0][11]            
            estado = db.excute_query(f"""select estado from enlace.estados WHERE id='{idEsatdo}';""")

            db.disconnect()
            
            if query and query2 and query3:
                subtotal_oce = 0
                isFisica = query4[0][0]
                name = f'{query3[0][3]} {query3[0][4]} {query3[0][5]}' if isFisica else f'{query3[0][2]}'
                address = f'Calle {query3[0][6]} #{query3[0][7]} {query3[0][8]},  Colonia {query3[0][9]} <br> {query3[0][10]}, {estado[0][0]} {query3[0][12]} <br> RFC: <b>{query3[0][13]}</b>'
                for row in query:
                    subtotal_oce = row[8] + subtotal_oce
                partida1 = float(subtotal_oce) * float(comision)
                subtotal = float(partida1) + float(partida2)
                iva = float(subtotal) * 0.16
                total = float(subtotal) + iva
                now = datetime.datetime.now()
                date = now.strftime('%d-%m-%Y')
                context = {
                    "title":"Factura Proveedor",
                    "idUsuario":idUsuario,
                    "occ":occ,
                    "oce":oce,
                    "name":name,
                    "address":address,
                    "comision":comision_original,
                    "subtotal":round(subtotal, 2),
                    "iva":round(iva, 2),
                    "total":round(total, 2),
                    "partida1":partida1,
                    "partida2":partida2,
                    "date":date,
                }

            if not query:
                messages.success(request, f"OCC '{occ}' no encontrada.")
            if not query2:
                messages.success(request, f"Usuario '{idUsuario}' no encontrado.")

            return render(request, 'crm/facturacion_proveedor.html', context)
        
        except Exception as e:
            messages.error(request, f'Ha occurrido un error: {e}')
            
    context = {"title":"Factura Proveedor"}
    return render(request, 'crm/facturacion_proveedor.html', context)

def exportFacturacionCliente(request,occ,idUsuario):
    try:
        query = db.excute_query(f"""select  CONCAT( 'OCC',SOL.FOLIO) OCC_FOLIO, USR.username Proveedor,
                                SOD.partida,
                                COD.descripcion, COD.marca, COD.empaque, COD.cantidad, COD.precioUnitario, COD.precioExtendido, COD.iva
                                from enlace.solicitudes SOL 
                                inner join enlace.solicituddetalle SOD on SOL.id = SOD.idSolicitud
                                inner join enlace.cotizaciondetalle COD on COD.ID = SOD.idCotizacionDetalle
                                left JOIN enlace.USUARIOS USR ON USR.ID = COD.IDUSUARIO
                                WHERE  SOL.status = 'confirmada'
                                and CONCAT('OCC' , SOL.folio) = '{occ}';""")
        query2 = db.excute_query(f"""select * from enlace.usuarios 
                                    WHERE id='{idUsuario}';""")
        query3 = db.excute_query(f"""select * from enlace.datosfiscales
                                    WHERE idUsuario='{idUsuario}';""")
        
        # PERSONA FISICA  isPersonaFisica 1=Si, 0=No
        query4 = db.excute_query(f"""select isPersonaFisica from enlace.usuarios WHERE id='{idUsuario}';;""")

        query5 = db.excute_query(f"""select CONCAT( 'OCC',SOL.FOLIO) OCC_FOLIO, FL.alto, FL.ancho, FL.largo, FL.peso, FL.costo, FL.combustible, FL.IVA from enlace.fletesolicitud FL
                            inner join enlace.solicitudes SOL on SOL.id = FL.idSolicitud
                            where CONCAT('OCC' , SOL.folio) = '{occ}';'""")
        idEsatdo = query3[0][11]
        estado = db.excute_query(f"""select estado from enlace.estados WHERE id='{idEsatdo}';""")
        db.disconnect()
        
        if query and query2 and query3:
            subtotal = 0
            isFisica = query4[0][0]
            name = f'{query3[0][3]} {query3[0][4]} {query3[0][5]}' if isFisica else f'{query3[0][2]}'
            address = f'Calle {query3[0][6]} #{query3[0][7]} {query3[0][8]},  Colonia {query3[0][9]} <br> {query3[0][10]}, {estado[0][0]} {query3[0][12]} <br> RFC: <b>{query3[0][13]}</b>'
            costo_envio = query5[0][6]
            iva_envio = query5[0][7]
            for row in query:
                subtotal = row[8] + subtotal
            subtotal = float(subtotal) + float(costo_envio)
            iva = float(subtotal) * 0.16 
            total = float(subtotal) + iva
            
            #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            #Titulo
            ws['F3'] = 'PROVEEDORA OMNIPET S. DE R.L DE C.V'
            ws['F3'].font = Font(bold=True)
            ws.merge_cells('F3:J3')
            ws['F4'] = 'Av. Cuauhtemoc 451-209, Col. Narvarte,'
            ws.merge_cells('F4:J4')
            ws['F5'] = 'México, D.F. 03020'
            ws.merge_cells('F5:G5')
            ws['F6'] = 'POM 020313 MS0'
            ws.merge_cells('F6:G6')

            #datos fiscales del cliente
            ws['B8'] = f'{name}'
            ws['B8'].font = Font(bold=True)
            ws.merge_cells('B8:G8')
            ws['L8'] = 'México D.F'
            ws.merge_cells('L8:M8')
            now = datetime.datetime.now()
            date = now.strftime('%d-%m-%Y')
            ws['L9'] = f'{date}'
            ws.merge_cells('L9:N9')
            ws['L10'] = f'{occ}'
            ws['L10'].font = Font(bold=True)
            ws.merge_cells('L10:N10')
            ws['B9'] = f'Calle {query3[0][6]} #{query3[0][7]} {query3[0][8]},  Colonia {query3[0][9]}'
            ws.merge_cells('B9:G9')
            ws['B10'] = f'{query3[0][10]}, {estado[0][0]} {query3[0][12]}'
            ws.merge_cells('B10:F10')
            ws['B12'] = 'CLIENTE'
            ws['B12'].font = Font(bold=True)
            ws['B13'] = f'RFC: {query3[0][13]}'
            ws['B13'].font = Font(bold=True)
            ws.merge_cells('B13:D13')

            #datos de la factura
            ws['B17'] = 'Partida'
            ws['C17'] = 'Descripción'
            ws.merge_cells('C17:H17')
            ws['L17'] = 'Cantidad'
            ws['M17'] = 'P.Unitario'
            ws['N17'] = 'Total'
            row = 19

            for partida in query:
                ws[f'B{row}'] = f'{partida[2]}'
                ws[f'C{row}'] = f'{partida[3]} {partida[4]} {partida[5]}'
                ws.merge_cells(f'C{row}:K{row}')
                ws[f'L{row}'] = f'{partida[6]}'
                ws[f'M{row}'] = f'{partida[7]}'
                ws[f'N{row}'] = f'{partida[8]}'
                row = row + 1
            
            ws[f'C{row + 1}'] = f'COSTO DE ENVÍO'
            ws.merge_cells(f'C{row + 1}:K{row + 1}')
            ws[f'L{row + 1}'] = f'1'
            ws[f'M{row + 1}'] = f'{costo_envio}'
            ws[f'N{row + 1}'] = f'{iva_envio}'

            #Totales
            subtotal = round(subtotal, 2)
            iva = round(iva, 2)
            total = round(total, 2)
            ws[f'L{row + 4}'] = 'Subtotal'
            ws[f'N{row + 4}'] = f'${subtotal}'
            ws[f'L{row + 5}'] = 'IVA'
            ws[f'N{row + 5}'] = f'${iva}'
            ws[f'L{row + 6}'] = 'Total'
            ws[f'N{row + 6}'] = f'${total}'
    
            #Guardamos el archivo
            nombre_archivo =f"{occ}.xlsx"
            response = HttpResponse(content_type="application/ms-excel") 
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response

        if not query:
            messages.success(request, f"OCC '{occ}' no encontrada.")
        if not query2:
            messages.success(request, f"Usuario '{idUsuario}' no encontrado.")
    
    except Exception as e:
        messages.error(request, f'Ha occurrido un error: {e}')

def exportFacturacionProveedor(request,occ,oce,idUsuario,comision,envio):
    try:
        query = db.excute_query(f"""select CONCAT('OCC' , SOL.folio) OCC,
                                        (select CONCAT( 'OCE', CO.FOLIO)  from enlace.cotizacion CO where CONCAT( 'OCE', CO.FOLIO) = '{oce}') OCE,
                                        SOD.partida, COD.descripcion, COD.marca, COD.empaque, COD.cantidad, COD.precioUnitario, COD.precioExtendido, COD.iva
                                        from enlace.solicitudes SOL 
                                        inner join enlace.solicituddetalle SOD on SOL.id = SOD.idSolicitud
                                        inner join enlace.cotizaciondetalle COD on COD.ID = SOD.idCotizacionDetalle
                                        inner JOIN enlace.USUARIOS USR ON USR.ID = COD.IDUSUARIO
                                        WHERE  USR.id in (select idUsuario from enlace.cotizacion CO where CONCAT( 'OCE', CO.FOLIO) = '{oce}') 
                                        and CONCAT('OCC' , SOL.folio) = '{occ}';""")
        query2 = db.excute_query(f"""select * from enlace.usuarios 
                                    WHERE id='{idUsuario}';""")
        query3 = db.excute_query(f"""select * from enlace.datosfiscales
                                    WHERE idUsuario='{idUsuario}';""")
        
        # PERSONA FISICA  isPersonaFisica 1=Si, 0=No
        query4 = db.excute_query(f"""select isPersonaFisica from enlace.usuarios WHERE id='{idUsuario}';;""")
        
        # ESTADO
        idEsatdo = query3[0][11]            
        estado = db.excute_query(f"""select estado from enlace.estados WHERE id='{idEsatdo}';""")

        db.disconnect()
        
        if query and query2 and query3:
            subtotal_oce = 0
            comision = float(comision) * 0.01
            isFisica = query4[0][0]
            name = f'{query3[0][3]} {query3[0][4]} {query3[0][5]}' if isFisica else f'{query3[0][2]}'
            address = f'Calle {query3[0][6]} #{query3[0][7]} {query3[0][8]},  Colonia {query3[0][9]} <br> {query3[0][10]}, {estado[0][0]} {query3[0][12]} <br> RFC: <b>{query3[0][13]}</b>'
            for row in query:
                subtotal_oce = row[8] + subtotal_oce
            partida1 = float(subtotal_oce) * float(comision)
            subtotal = float(partida1) + float(envio)
            iva = float(subtotal) * 0.16
            total = float(subtotal) + iva
            now = datetime.datetime.now()
            date = now.strftime('%d-%m-%Y')
            
            #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active
            #Titulo
            ws['F3'] = 'PROVEEDORA OMNIPET S. DE R.L DE C.V'
            ws['F3'].font = Font(bold=True)
            ws.merge_cells('F3:J3')
            ws['F4'] = 'Av. Cuauhtemoc 451-209, Col. Narvarte,'
            ws.merge_cells('F4:J4')
            ws['F5'] = 'México, D.F. 03020'
            ws.merge_cells('F5:G5')
            ws['F6'] = 'POM 020313 MS0'
            ws.merge_cells('F6:G6')

            #datos fiscales del cliente
            ws['B8'] = f'{name}'
            ws['B8'].font = Font(bold=True)
            ws.merge_cells('B8:G8')
            ws['J8'] = 'EXPEDIDO EN:'
            ws.merge_cells('J8:K8')
            ws['L8'] = 'México D.F'
            ws.merge_cells('L8:M8')
            ws['K9'] = 'FECHA:'
            now = datetime.datetime.now()
            date = now.strftime('%d-%m-%Y')
            ws['L9'] = f'{date}'
            ws.merge_cells('L9:N9')
            # ws['L10'] = f'{occ}'
            # ws['L10'].font = Font(bold=True)
            # ws.merge_cells('L10:N10')
            ws['B9'] = f'Calle {query3[0][6]} #{query3[0][7]} {query3[0][8]},  Colonia {query3[0][9]}'
            ws.merge_cells('B9:I9')
            ws['B10'] = f'{query3[0][10]}, {estado[0][0]} {query3[0][12]}'
            ws.merge_cells('B10:F10')
            ws['B11'] = f'RFC: {query3[0][13]}'
            ws['B11'].font = Font(bold=True)
            ws.merge_cells('B11:D11')

            #datos de la factura
            ws['B17'] = 'Partida'
            ws['C17'] = 'Descripción'
            ws.merge_cells('C17:H17')
            ws['L17'] = 'Cantidad'
            ws['M17'] = 'P.Unitario'
            ws['N17'] = 'Total'

            ws['B19'] = '1'      
            ws['C19'] = 'COMISION POR USO DE LA PLATAFORMA MIENLACE'
            ws.merge_cells('C19:K19')
            ws['L19'] = '1'
            ws['M19'] = f'$ {partida1}'
            ws['N19'] = f'$ {partida1}'
            
            ws['B21'] = '2'    
            ws['C21'] = 'COSTO DE ENVÍO'
            ws.merge_cells('C21:K21')
            ws['L21'] = '1'
            ws['M21'] = f'$ {envio}'
            ws['N21'] = f'$ {envio}'

            #Totales
            subtotal = round(subtotal, 2)
            iva = round(iva, 2)
            total = round(total, 2)
            ws[f'L27'] = 'Subtotal'
            ws[f'N27'] = f'${subtotal}'
            ws[f'L28'] = 'IVA'
            ws[f'N28'] = f'${iva}'
            ws[f'L29'] = 'Total'
            ws[f'N29'] = f'${total}'

            #Bordes variables
            double = Side(border_style="double", color="000000")

            #Formato de celdas
            for c in ws['B17:N17'][0]:
                c.border = Border(bottom=double, top=double)
                c.font = Font(bold=True)
            
            for c in ws['B19:N19'][0]:
                c.fill = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')
            
            for c in ws['B21:N21'][0]:
                c.fill = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')
            
            for c in ws['L27:N27'][0]:
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')

            for c in ws['L28:N28'][0]:
                c.font = Font(bold=True)
            
            for c in ws['L29:N29'][0]:
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')

            #Guardamos el archivo
            nombre_archivo =f"{oce}-{idUsuario}.xlsx"
            response = HttpResponse(content_type="application/ms-excel") 
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response

        if not query:
            messages.success(request, f"OCC '{occ}' no encontrada.")
        if not query2:
            messages.success(request, f"Usuario '{idUsuario}' no encontrado.")
    
    except Exception as e:
        messages.error(request, f'Ha occurrido un error: {e}')
            
    